"""Parse XLSX and CSV files into detected-section dicts for the import pipeline."""
from __future__ import annotations

import csv
import io
from typing import Any

from app.imports.mapping_detector import (
    ColumnMapping,
    detect_column_mapping,
    detect_section_type,
    is_sample_content,
    score_row,
)
from app.schemas.imports import DetectedRow, DetectedSection


def _rows_to_detected_section(
    rows: list[dict[str, Any]],
    headers: list[str],
    sheet_name: str | None,
    col_map: ColumnMapping,
) -> DetectedSection:
    first_values = [str(v) for v in list(rows[0].values())[:5]] if rows else []
    section_type = detect_section_type(headers, first_values)
    is_sample, sample_reason = is_sample_content(
        sheet_name, [str(r.get(col_map.name_col, "")) for r in rows[:5] if col_map.name_col]
    )

    detected_rows: list[DetectedRow] = []
    for i, row in enumerate(rows):
        # Skip completely empty rows
        if all(v is None or str(v).strip() == "" for v in row.values()):
            continue
        mapped, row_conf, flags = score_row(row, col_map)
        if mapped.get("name"):  # only keep rows with at least a name
            detected_rows.append(
                DetectedRow(
                    row_index=i,
                    raw={k: str(v) if v is not None else "" for k, v in row.items()},
                    mapped=mapped,
                    confidence=row_conf,
                    flags=flags,
                )
            )

    section_conf = col_map.confidence * (0.9 if is_sample else 1.0)
    return DetectedSection(
        section_type=section_type,
        sheet_name=sheet_name,
        rows=detected_rows,
        confidence=round(section_conf, 3),
        is_sample=is_sample,
        sample_reason=sample_reason,
    )


def parse_xlsx(file_bytes: bytes) -> list[DetectedSection]:
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    sections: list[DetectedSection] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            continue

        # First non-empty row is headers
        header_row = None
        data_start = 0
        for idx, row in enumerate(all_rows):
            non_empty = [c for c in row if c is not None and str(c).strip()]
            if len(non_empty) >= 2:
                header_row = [str(c).strip() if c else f"col_{i}" for i, c in enumerate(row)]
                data_start = idx + 1
                break

        if header_row is None:
            continue

        col_map = detect_column_mapping(header_row)
        rows_as_dicts: list[dict] = []
        for row in all_rows[data_start:]:
            rows_as_dicts.append({header_row[i]: row[i] for i in range(min(len(header_row), len(row)))})

        if not rows_as_dicts:
            continue

        section = _rows_to_detected_section(rows_as_dicts, header_row, sheet_name, col_map)
        sections.append(section)

    return sections


def parse_csv(file_bytes: bytes) -> list[DetectedSection]:
    text = file_bytes.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    headers = reader.fieldnames or []
    rows = list(reader)

    if not headers or not rows:
        return []

    col_map = detect_column_mapping(list(headers))
    section = _rows_to_detected_section(
        [{k: v for k, v in row.items()} for row in rows],
        list(headers),
        None,
        col_map,
    )
    return [section]
